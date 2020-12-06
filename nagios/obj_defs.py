# obj_defs.py - web scrap nagios "object definitions" doc page 

import sys
import json
from lxml import etree, html
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#-------------------------------------------------------------------------------
# get_text
#-------------------------------------------------------------------------------

def get_text(p):
    """Extract all the text under this node"""
    s = p.text.rstrip() + ' '
    for k in p:
        s += k.text if k.text is not None else ''
        s += k.tail if k.tail is not None else ''

    s = s.replace('\n', '')
    return s

#-------------------------------------------------------------------------------
# get_directives
#-------------------------------------------------------------------------------

def get_directives(tbl):
    d = []

    first_buggy_td = True
    
    # Each row has two td's, name and description
    for row in tbl:
        # There's an empty row after the "importance" directive
        if len(row) == 0:
            continue

        # Bug in service object, event_handler directive: missing <tr> tag, the
        # two <td> nodes appear as children of the table. I'm assuming this is
        # the onlly bug, and that the two <td> elements come one after the
        # other.
        if row.tag == 'td':
            if first_buggy_td:
                name = row[0].text
                # print(f'    {name}')
                first_buggy_td = False
                continue
            else:
                # Second column is the description
                description = get_text(row)
                d.append((name, description))
                first_buggy_td = True
        
        # First column is the attribute name (inside a span)
        td = row[0]
        # s = td.text if td.text is not None else ''
        # print(f'len(td.attrib)={len(td.attrib)}, len(td)={len(td)}, td.text="{s}"')
        if len(td) > 0:
            name = td[0].text
            # print(f'    {name}')
            # Second column is the description 
            description = get_text(row[1])
            d.append((name, description))

    return d

#-------------------------------------------------------------------------------
# get_main
#-------------------------------------------------------------------------------

def get_main(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        root = html.fromstring(f.read())

    # Find the right section
    p = root.find('.//div[@class="col-sm-12"]')

    # p is a <div> element. The first 11 children are specific, after that we
    # have a repeating structure:
    #    
    #    <p> <a name="xxx">...
    #    <table>...
    #    <p class="SectionTitle">Description:</p>
    #    <p>
    #    <p class="SectionTitle">Definition Format:</p>
    #    <p>
    #    <table>...
    #    <p class="SectionTitle">Example Definition:</p>
    #    <pre>...
    #    <p class="SectionTitle">Directive Descriptions:</p>
    #    <table>...

    objects = []
    state = None
    for k in p[11:]:
        if k.tag == 'p':
            if state == None:
                obj = {'name': k[0].attrib['name']}
                # print(obj['name'])
                state = 'start'
            elif state == 'header':
                state = 'title_desc'
            elif state == 'title_desc':
                obj['description'] = k.text.strip()
                state = 'description'
            elif state == 'description':
                state = 'title_format'
            elif state == 'title_format':
                state = 'format'
            elif state == 'format_table':
                state = 'example_def'
            elif state == 'example_def':
                state = 'title_directive'
                
        elif k.tag == 'table':
            if state == 'start':
                x = k.find('.//tr/td[@class="SectionHeader"]')
                obj['header'] = x.text
                state = 'header'
            elif state == 'format':
                state = 'format_table'
            elif state == 'title_directive':
                obj['directives']= get_directives(k)
                objects.append(obj)
                state = None

        elif k.tag == 'pre':
            continue
        
    return objects 

#-------------------------------------------------------------------------------
# dump_objects
#-------------------------------------------------------------------------------

def dump_objects(objects):
    s = ''
    ind = ' '*4
    for obj in objects:
        s += f'name: {obj["name"]}\n'
        s += f'{ind}header: {obj["header"]}\n'
        s += f'{ind}description: {obj["description"]}\n'
        s += f'{ind}directives:\n'
        for dir in obj['directives']:
            s += f'{ind*2}{dir[0]}\n'
            s += f'{ind*2}{dir[1]}\n'
    s += '\n'
    return s

#-------------------------------------------------------------------------------
# dump_ws
#-------------------------------------------------------------------------------

def dump_ws(ws, obj):
    # Worksheet title
    ws.title = obj['name']
    
    # Header
    ws.cell(row=1, column=1, value='Directive')
    ws.cell(row=1, column=2, value='Description')

    row = 2
    for dir in obj['directives']:
        ws.cell(row=row, column=1, value=dir[0])
        ws.cell(row=row, column=2, value=dir[1])
        row += 1

    # Object header and description
    ws.cell(row=row, column=1, value='Header')
    ws.cell(row=row, column=2, value=obj['header'])
    row += 1
    ws.cell(row=row, column=1, value='Description')
    ws.cell(row=row, column=2, value=obj['description'])

#-------------------------------------------------------------------------------
# dump_excel
#-------------------------------------------------------------------------------

def dump_excel(objects):
    wb = Workbook()
    ws = wb.active
    dump_ws(ws, objects[0])

    for obj in objects[1:]:
        ws = wb.create_sheet()
        dump_ws(ws, obj)

    wb.save(filename='obj_defs.xlsx')

#-------------------------------------------------------------------------------
# main
#-------------------------------------------------------------------------------

if len(sys.argv) != 2:
    print(f'Usage: {sys.argv[0]} <filepath>')
    sys.exit(1)

filepath = sys.argv[1]
objects = get_main(filepath)
# print(dump_objects(objects))
dump_excel(objects)