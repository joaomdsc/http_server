# main_cfg.py - web scrap nagios "main config file" doc page 

import sys
import json
from lxml import etree, html
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#-------------------------------------------------------------------------------
# get_title
#-------------------------------------------------------------------------------

def get_title(tbl):
    x = tbl.find('.//tr/td/strong')
    return x.text if x is not None else ''

#-------------------------------------------------------------------------------
# get_examples
#-------------------------------------------------------------------------------

def get_examples(tbl):
    # Each row has two td's, the first is known (Format or Example)
    d = {}

    # First row is the format. First td is the "Format:" label, we want the 2nd.
    row = tbl[0]
    td = row[1].find('.//strong')
    s = td.text
    s = s.split('=', maxsplit=1)[1]
    s = s[1:-1]
    d['format'] = s

    # Second row is the examples table. First td is the "Example:" label, we want the 2nd. 
    row = tbl[1]
    examples = []
    for fn in row[1].findall('.//font[@color="red"]/strong'):
        examples.append(fn.text)
    d['examples'] = examples

    return d

#-------------------------------------------------------------------------------
# get_text
#-------------------------------------------------------------------------------

def get_text(p):
    s = p.text.rstrip() + ' '
    for k in p:
        s += k.text
        s += k.tail if k.tail is not None else ''

    s = s.replace('\n', '')
    return s

#-------------------------------------------------------------------------------
# get_main
#-------------------------------------------------------------------------------

def get_main(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        root = html.fromstring(f.read())

    # Find the right section
    for x in root.findall('.//div[@class="section"]/h4'):
        if x.text == 'Configuration File Variables':
            break
    # print(x.text)
    p = x.getparent()

    # p is a <div class="section"> element. The first two children are h4 and
    # p, after that we have a repeating structure:
    #    <a name="xxx">
    #    <table> ...
    #    <br>
    #    <table> ...
    #    <p> ...

    options = []
    state = None
    for k in p[2:]:
        if k.tag == 'a':
            opt = {
                'name': k.attrib['name']
            }
            # print(opt['name'])
            state = 'opt_start'
        elif k.tag == 'table':
            if state == 'opt_start':
                opt['title'] = get_title(k)
                state = 'opt_title'
            elif state == 'opt_title': 
                opt['fmt_ex'] = get_examples(k)
                state = 'opt_fmt_ex'
        elif k.tag == 'p':
               opt['text'] = get_text(k)
               options.append(opt)

    return options 

#-------------------------------------------------------------------------------
# dump_options
#-------------------------------------------------------------------------------

def dump_options(options):
    s = ''
    ind = ' '*4
    for opt in options:
        for k, v in opt.items():
            if k == 'name':
                s += f'{k}: {v}\n'
            elif k == 'title':
                s += f'{ind}{k}: {v}\n'
            elif k == 'fmt_ex':
                # v is a dict with keys format, examples
                s += f'{ind}format: {v["format"]}\n'
                s += f'{ind}examples:\n'
                for e in v['examples']:
                    s += f'{ind*2}{e}\n'
            elif k == 'text':
                s += v

    s += '\n'
    return s

#-------------------------------------------------------------------------------
# dump_excel
#-------------------------------------------------------------------------------

def dump_excel(options):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Main Cfg"

    # Header
    ws1.cell(row=1, column=1, value='Title')
    ws1.cell(row=1, column=2, value='Name')
    ws1.cell(row=1, column=3, value='Format')
    ws1.cell(row=1, column=4, value='Description')

    row = 2
    for opt in options:
        ws1.cell(row=row, column=1, value=opt['title'])
        ws1.cell(row=row, column=2, value=opt['name'])
        ws1.cell(row=row, column=3, value=opt['fmt_ex']['format'])
        ws1.cell(row=row, column=4, value=opt['text'])
        row += 1

    wb.save(filename='main_cfg.xlsx')

#-------------------------------------------------------------------------------
# main
#-------------------------------------------------------------------------------

if len(sys.argv) != 2:
    print(f'Usage: {sys.argv[0]} <filepath>')
    sys.exit(1)

filepath = sys.argv[1]
options = get_main(filepath)
# print(dump_options(options))
dump_excel(options)